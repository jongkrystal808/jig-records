import csv
from io import BytesIO, StringIO

from openpyxl import load_workbook

from backend.app.services.configuration_report_service import ConfigurationReportService
from backend.app.services.inventory_service import InventoryService
from backend.app.utils.csv_tools import escape_spreadsheet_formula, render_csv_text, stream_csv_text


def test_csv_renderers_escape_formula_prefixes_only_for_strings() -> None:
    dangerous = ["=1+1", "+SUM(A1:A2)", "-2+3", "@cmd"]
    rows = [{"text": value, "quantity": -5} for value in dangerous]

    rendered = list(csv.DictReader(StringIO(render_csv_text(["text", "quantity"], rows))))
    streamed = list(csv.DictReader(StringIO("".join(stream_csv_text(["text", "quantity"], iter(rows))).lstrip("\ufeff"))))

    assert [row["text"] for row in rendered] == [f"'{value}" for value in dangerous]
    assert [row["text"] for row in streamed] == [f"'{value}" for value in dangerous]
    assert [row["quantity"] for row in rendered] == ["-5"] * len(dangerous)
    assert escape_spreadsheet_formula(-5) == -5
    assert escape_spreadsheet_formula("ordinary") == "ordinary"


def test_configuration_report_csv_and_xlsx_escape_formula_cells() -> None:
    headers = ["Name", "Quantity"]
    rows = [["=HYPERLINK(\"bad\")", -3]]

    csv_rows = list(csv.reader(StringIO(ConfigurationReportService.render_csv(headers, rows).decode("utf-8-sig"))))
    assert csv_rows[1] == ["'=HYPERLINK(\"bad\")", "-3"]

    workbook = load_workbook(BytesIO(ConfigurationReportService.render_xlsx(headers, rows)), data_only=False)
    worksheet = workbook.active
    assert worksheet["A2"].value == "'=HYPERLINK(\"bad\")"
    assert worksheet["A2"].data_type == "s"
    assert worksheet["B2"].value == -3


def test_transaction_xlsx_escapes_formula_cells() -> None:
    content = InventoryService.render_transaction_report_xlsx(
        "Transaction report",
        ["note", "quantity"],
        [{"note": "@malicious", "quantity": -8}],
    )
    worksheet = load_workbook(BytesIO(content), data_only=False).active
    assert worksheet["A3"].value == "'@malicious"
    assert worksheet["A3"].data_type == "s"
    assert worksheet["B3"].value == -8
